# @Time : 2020/3/8 22:37
# @Author : shell-craw
# @File : pyredis.py
import openpyxl
from openpyxl import Workbook
from collections import Counter
import time


wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Sheet2']
# ws1 = wb.get_sheet_names()
# ce = ws.cell(row=1, column=1)
# print(ce.value)
# sn = ws.cell(row=2, column=14)
# print(sn.value)
# rows_data = list(ws.rows)[1:]
# for case in rows_data:
    # print(case)
    # cases = []
    # for cell in case:
        # pass
        # cases.append(cell.value)
    # print(cases)

# 获取SN数据
# 求最大行数
max_r = ws.max_row
case_data = []
# 遍历SN列
for row in range(1, max_r):
    if row != 1:
        # 将SN的数据赋值给info
        info = ws.cell(row, 2).value
        # SN数据保存到列表中
        case_data.append(info)
# print(case_data)

# 获取重复的元素
repeat_data = []
repeat_SN = {}
# 统计每个SN码出现的次数，保存到字典中，SN为键，出现次数为值
repeat = dict(Counter(case_data))
# print(repeat)
# 获取字典的键和值
for k, v in repeat.items():
    # 如果出现次数大于1保存到repeat_data中
    if v > 1:
        repeat_data.append(k)
        repeat_SN[k] = v
        # print(k)
# print(repeat_SN)

# 存储相同SN码的索引
re_sn = []
# 将re_sn存储到一起
merge = []
# 获取重复元素值及下标
for ke, va in repeat_SN.items():
    # enumerate可以获取每个元素的下标，并保存到元组中(0, 'SN')
    for i, x in enumerate(case_data):
        if x == ke:
            re_sn.append(i+2)
            if len(re_sn) >= va:
                merge.append(re_sn)
                re_sn = []
            # print(i+2)
# print(merge)

# 去重
store_li = []
time_li = []
for midx in merge:
    for i in range(len(midx)):
        store_li.append(ws.cell(row=midx[i], column=1).value)
        time_li.append(ws.cell(row=midx[i], column=3).value)
    print(time_li)
    for i in range(1, len(time_li)):
        if time_li[0] < time_li[i] and store_li[0] == store_li[i]:
            ws['D{}'.format(midx[i])] = 'delete'
        elif time_li[0] == time_li[i] and store_li[0] == store_li[i]:
            ws['D{}'.format(midx[i])] = 'delete'

    min_time = min(time_li)
    m = time_li.index(min_time)
    time_li = []
# print(time_li)
        # print(store, time, '')

wb.save('test.xlsx')
wb.close()